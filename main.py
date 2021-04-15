import random
import xml.etree.ElementTree as ET

import openpyxl

from HeapSort import HeapSort

tree = ET.parse("data.xml")
wb = openpyxl.load_workbook('result.xlsx')
sheet = wb['MassTask']
sheet['A1'] = 'Длина массива'
sheet['B1'] = 'Кол-во итераций'
root = tree.getroot()
j = 2
for child in root:
    start_length = int(child.attrib['startLength'])
    max_length = int(child.attrib['maxLength'])
    min_element = int(child.attrib['minElement'])
    max_element = int(child.attrib['maxElement'])
    diff = float(child.attrib['diff'])
    while start_length <= max_length:
        for repeat in range(int(child.attrib['repeat'])):
            arr = []
            for i in range(start_length):
                arr.append(random.randint(min_element, max_element))
            heap_sort = HeapSort()
            heap_sort.sort(arr)
            sheet.cell(row=j, column=1).value = len(arr)
            sheet.cell(row=j, column=2).value = heap_sort.iteration_counter
            j += 1
        if child.attrib['name'] == 'Arithmetic Progression':
            start_length = int(start_length + diff)
        else:
            start_length = int(start_length * diff)

wb.save('result.xlsx')
